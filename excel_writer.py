import os
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell import MergedCell

# 定数設定
TEMPLATE_PATH = "template.xlsx"
OUTPUT_DIR = "output"

COLUMN_TO_CELL_COORDINATE_MAP = {
    # 主要なテキスト項目 (AI生成)
    "name": ("様式23_1", "F3"),
    "age": ("様式23_1", "AC3"),
    "header_disease_name_txt": ("様式23_1", "B5"),
    "main_comorbidities_txt": ("様式23_1", "B8"),
    "main_risks_txt": ("様式23_1", "R8"),
    "main_contraindications_txt": ("様式23_1", "AH8"),
    "func_pain_txt": ("様式23_1", "H25"),
    "func_muscle_weakness_txt": ("様式23_1", "AF14"),
    "adl_equipment_and_assistance_details_txt": ("様式23_1", "AE35"),
    "goals_1_month_txt": ("様式23_1", "B74"),
    "goals_at_discharge_txt": ("様式23_1", "Z74"),
    "policy_treatment_txt": ("様式23_1", "B79"),
    "policy_content_txt": ("様式23_1", "Z79"),
    "goal_p_action_plan_txt": ("様式23_2", "Z2"),
    "goal_a_action_plan_txt": ("様式23_2", "Z13"),
    "goal_s_psychological_action_plan_txt": ("様式23_2", "Z52"),
    "goal_s_env_action_plan_txt": ("様式23_2", "Z55"),
    "goal_s_3rd_party_action_plan_txt": ("様式23_2", "Z65"),
    "func_swallowing_disorder_txt": ("様式23_1", "H21"),
    "func_behavioral_psychiatric_disorder_txt": ("様式23_1", "AF22"),
    "goal_p_household_role_txt": ("様式23_2", "J10"),
    "goal_p_hobby_txt": ("様式23_2", "J12"),
    "func_rom_limitation_txt": ("様式23_1", "AF12"),
    # 日付 (特殊処理用)
    "header_evaluation_date": ("様式23_1", "AN3"),
    "header_evaluation_year_txt": ("様式23_1", "AN3"),
    "header_evaluation_month_txt": ("様式23_1", "AQ3"),
    "header_evaluation_day_txt": ("様式23_1", "AT3"),
    "header_onset_date": ("様式23_1", "AN4"),
    "header_onset_year_txt": ("様式23_1", "AN4"),
    "header_onset_month_txt": ("様式23_1", "AQ4"),
    "header_onset_day_txt": ("様式23_1", "AT4"),
    "header_rehab_start_date": ("様式23_1", "AN5"),
    "header_rehab_start_year_txt": ("様式23_1", "AN5"),
    "header_rehab_start_month_txt": ("様式23_1", "AQ5"),
    "header_rehab_start_day_txt": ("様式23_1", "AT5"),
    "signature_explanation_date": ("様式23_1", "AP86"),
    "signature_explanation_year_txt": ("様式23_1", "AP86"),
    "signature_explanation_month_txt": ("様式23_1", "AS86"),
    "signature_explanation_day_txt": ("様式23_1", "AU86"),
    # 全座標リスト
    "func_motor_paralysis_chk": ("様式23_1", "AA16"),
    "func_speech_articulation_chk": ("様式23_1", "AA20"),
    "func_developmental_asd_chk": ("様式23_1", "AA26"),
    "nutrition_weight_val": ("様式23_1", "AA60"),
    "nutrition_method_tube_chk": ("様式23_1", "AA61"),
    "adl_eating_bi_current_val": ("様式23_1", "AB35"),
    "adl_grooming_bi_current_val": ("様式23_1", "AB36"),
    "adl_bathing_bi_current_val": ("様式23_1", "AB37"),
    "adl_dressing_bi_current_val": ("様式23_1", "AB38"),
    "adl_toileting_bi_current_val": ("様式23_1", "AB40"),
    "adl_bladder_management_bi_current_val": ("様式23_1", "AB41"),
    "adl_bowel_management_bi_current_val": ("様式23_1", "AB42"),
    "adl_transfer_bi_current_val": ("様式23_1", "AB43"),
    "adl_locomotion_walk_walkingAids_wc_bi_current_val": ("様式23_1", "AB46"),
    "adl_locomotion_stairs_bi_current_val": ("様式23_1", "AB49"),
    "header_therapy_pt_chk": ("様式23_1", "S5"),
    "header_therapy_ot_chk": ("様式23_1", "X5"),
    "header_therapy_st_chk": ("様式23_1", "AC5"),
    "nutrition_status_assessment_overnutrition_chk": ("様式23_1", "AC63"),
    "func_speech_aphasia_chk": ("様式23_1", "AD20"),
    "social_disability_certificate_mental_chk": ("様式23_1", "AD70"),
    "social_disability_certificate_mental_rank_val": ("様式23_1", "AD72"),
    "func_basic_sitting_balance_independent_chk": ("様式23_1", "AE29"),
    "func_basic_standing_balance_independent_chk": ("様式23_1", "AE30"),
    "func_basic_other_txt": ("様式23_1", "AE31"),
    "nutrition_method_iv_chk": ("様式23_1", "AE61"),
    "func_contracture_deformity_txt": ("様式23_1", "AF13"),
    "func_motor_involuntary_movement_chk": ("様式23_1", "AF16"),
    "func_motor_muscle_tone_abnormality_txt": ("様式23_1", "AF17"),
    "func_sensory_hearing_chk": ("様式23_1", "AF18"),
    "func_higher_brain_memory_chk": ("様式23_1", "AF21"),
    "func_disorientation_txt": ("様式23_1", "AF23"),
    "func_memory_disorder_txt": ("様式23_1", "AF24"),
    "func_speech_stuttering_chk": ("様式23_1", "AG20"),
    "func_developmental_ld_chk": ("様式23_1", "AH26"),
    "func_basic_sitting_balance_partial_assistance_chk": ("様式23_1", "AH29"),
    "func_basic_standing_balance_partial_assistance_chk": ("様式23_1", "AH30"),
    "func_sensory_vision_chk": ("様式23_1", "AI18"),
    "func_higher_brain_attention_chk": ("様式23_1", "AI21"),
    "nutrition_bmi_chk": ("様式23_1", "AI60"),
    "signature_explained_to_txt": ("様式23_1", "AI86"),
    "func_speech_other_chk": ("様式23_1", "AJ20"),
    "nutrition_method_iv_peripheral_chk": ("様式23_1", "AJ61"),
    "nutrition_status_assessment_other_chk": ("様式23_1", "AJ63"),
    "social_disability_certificate_intellectual_chk": ("様式23_1", "AK70"),
    "social_disability_certificate_intellectual_txt": ("様式23_1", "AK71"),
    "social_disability_certificate_intellectual_grade_txt": ("様式23_1", "AK72"),
    "func_motor_ataxia_chk": ("様式23_1", "AL16"),
    "func_sensory_superficial_chk": ("様式23_1", "AL18"),
    "func_higher_brain_apraxia_chk": ("様式23_1", "AL21"),
    "func_basic_sitting_balance_assistance_chk": ("様式23_1", "AL29"),
    "func_basic_standing_balance_assistance_chk": ("様式23_1", "AL30"),
    "func_developmental_adhd_chk": ("様式23_1", "AM26"),
    "nutrition_method_iv_central_chk": ("様式23_1", "AM61"),
    "nutrition_status_assessment_other_txt": ("様式23_1", "AN63"),
    "nutrition_required_protein_val": ("様式23_1", "AN65"),
    "nutrition_total_intake_protein_val": ("様式23_1", "AN66"),
    "goals_planned_hospitalization_period_chk": ("様式23_1", "AN73"),
    "goals_discharge_destination_chk": ("様式23_1", "AN74"),
    "goals_long_term_care_needed_chk": ("様式23_1", "AN75"),
    "func_speech_other_txt": ("様式23_1", "AO20"),
    "func_higher_brain_agnosia_chk": ("様式23_1", "AO21"),
    "func_basic_sitting_balance_not_performed_chk": ("様式23_1", "AO29"),
    "func_basic_standing_balance_not_performed_chk": ("様式23_1", "AO30"),
    "func_sensory_deep_chk": ("様式23_1", "AP18"),
    "nutrition_bmi_val": ("様式23_1", "AP60"),
    "func_motor_parkinsonism_chk": ("様式23_1", "AQ16"),
    "nutrition_method_peg_chk": ("様式23_1", "AQ61"),
    "func_higher_brain_executive_chk": ("様式23_1", "AR21"),
    "social_disability_certificate_other_chk": ("様式23_1", "AR70"),
    "social_disability_certificate_other_txt": ("様式23_1", "AR71"),
    "goals_discharge_destination_txt": ("様式23_1", "AR74"),
    "goals_planned_hospitalization_period_txt": ("様式23_1", "AT73"),
    "func_swallowing_disorder_chk": ("様式23_1", "B21"),
    "func_nutritional_disorder_chk": ("様式23_1", "B22"),
    "func_excretory_disorder_chk": ("様式23_1", "B23"),
    "func_pressure_ulcer_chk": ("様式23_1", "B24"),
    "func_pain_chk": ("様式23_1", "B25"),
    "func_other_chk": ("様式23_1", "B26"),
    "func_basic_rolling_chk": ("様式23_1", "B29"),
    "func_basic_getting_up_chk": ("様式23_1", "B30"),
    "func_basic_standing_up_chk": ("様式23_1", "B31"),
    "social_care_level_status_chk": ("様式23_1", "B70"),
    "func_respiratory_disorder_chk": ("様式23_1", "C13"),
    "func_respiratory_o2_therapy_chk": ("様式23_1", "C14"),
    "func_circulatory_disorder_chk": ("様式23_1", "C15"),
    "func_circulatory_ef_chk": ("様式23_1", "C16"),
    "func_risk_factors_chk": ("様式23_1", "C17"),
    "func_risk_hypertension_chk": ("様式23_1", "C18"),
    "func_risk_obesity_chk": ("様式23_1", "C19"),
    "func_risk_angina_chk": ("様式23_1", "C20"),
    "social_care_level_applying_chk": ("様式23_1", "C71"),
    "social_care_level_care_slct": ("様式23_1", "C72"),
    "func_circulatory_ef_val": ("様式23_1", "G16"),
    "func_basic_rolling_independent_chk": ("様式23_1", "G29"),
    "func_basic_getting_up_independent_chk": ("様式23_1", "G30"),
    "func_basic_standing_up_independent_chk": ("様式23_1", "G31"),
    "social_care_level_support_chk": ("様式23_1", "G71"),
    "func_respiratory_o2_therapy_l_min_txt": ("様式23_1", "H14"),
    "func_risk_dyslipidemia_chk": ("様式23_1", "H18"),
    "func_risk_hyperuricemia_chk": ("様式23_1", "H19"),
    "func_risk_omi_chk": ("様式23_1", "H20"),
    "func_nutritional_disorder_txt": ("様式23_1", "H22"),
    "func_excretory_disorder_txt": ("様式23_1", "H23"),
    "func_pressure_ulcer_txt": ("様式23_1", "H24"),
    "func_other_txt": ("様式23_1", "H26"),
    "signature_rehab_doctor_txt": ("様式23_1", "H86"),
    "signature_pt_txt": ("様式23_1", "H87"),
    "signature_st_txt": ("様式23_1", "H88"),
    "signature_dietitian_txt": ("様式23_1", "H89"),
    "signature_explainer_txt": ("様式23_1", "H91"),
    "nutrition_height_chk": ("様式23_1", "I60"),
    "social_care_level_care_num1_slct": ("様式23_1", "I72"),
    "func_consciousness_disorder_jcs_gcs_txt": ("様式23_1", "J12"),
    "func_basic_rolling_partial_assistance_chk": ("様式23_1", "J29"),
    "func_basic_getting_up_partial_assistance_chk": ("様式23_1", "J30"),
    "func_basic_standing_up_partial_assistance_chk": ("様式23_1", "J31"),
    "nutrition_status_assessment_no_problem_chk": ("様式23_1", "J63"),
    "social_care_level_care_num2_slct": ("様式23_1", "K72"),
    "adl_eating_fim_start_val": ("様式23_1", "M35"),
    "adl_grooming_fim_start_val": ("様式23_1", "M36"),
    "adl_bathing_fim_start_val": ("様式23_1", "M37"),
    "adl_dressing_upper_fim_start_val": ("様式23_1", "M38"),
    "adl_dressing_lower_fim_start_val": ("様式23_1", "M39"),
    "adl_toileting_fim_start_val": ("様式23_1", "M40"),
    "adl_bladder_management_fim_start_val": ("様式23_1", "M41"),
    "adl_bowel_management_fim_start_val": ("様式23_1", "M42"),
    "adl_transfer_bed_chair_wc_fim_start_val": ("様式23_1", "M43"),
    "adl_transfer_toilet_fim_start_val": ("様式23_1", "M44"),
    "adl_transfer_tub_shower_fim_start_val": ("様式23_1", "M45"),
    "adl_locomotion_walk_walkingAids_wc_fim_start_val": ("様式23_1", "M46"),
    "adl_locomotion_stairs_fim_start_val": ("様式23_1", "M49"),
    "adl_comprehension_fim_start_val": ("様式23_1", "M51"),
    "adl_expression_fim_start_val": ("様式23_1", "M52"),
    "adl_social_interaction_fim_start_val": ("様式23_1", "M53"),
    "adl_problem_solving_fim_start_val": ("様式23_1", "M54"),
    "adl_memory_fim_start_val": ("様式23_1", "M55"),
    "nutrition_swallowing_diet_None_chk": ("様式23_1", "M62"),
    "social_care_level_support_num1_slct": ("様式23_1", "M71"),
    "social_care_level_care_num3_slct": ("様式23_1", "M72"),
    "func_basic_rolling_assistance_chk": ("様式23_1", "N29"),
    "func_basic_getting_up_assistance_chk": ("様式23_1", "N30"),
    "func_basic_standing_up_assistance_chk": ("様式23_1", "N31"),
    "func_respiratory_tracheostomy_chk": ("様式23_1", "O14"),
    "func_risk_diabetes_chk": ("様式23_1", "O18"),
    "func_risk_ckd_chk": ("様式23_1", "O19"),
    "func_risk_other_chk": ("様式23_1", "O20"),
    "nutrition_method_oral_chk": ("様式23_1", "O61"),
    "nutrition_swallowing_diet_True_chk": ("様式23_1", "O62"),
    "social_care_level_support_num2_slct": ("様式23_1", "O71"),
    "social_care_level_care_num4_slct": ("様式23_1", "O72"),
    "nutrition_height_val": ("様式23_1", "P60"),
    "nutrition_status_assessment_malnutrition_chk": ("様式23_1", "P63"),
    "func_basic_rolling_not_performed_chk": ("様式23_1", "Q29"),
    "func_basic_getting_up_not_performed_chk": ("様式23_1", "Q30"),
    "func_basic_standing_up_not_performed_chk": ("様式23_1", "Q31"),
    "social_care_level_care_num5_slct": ("様式23_1", "Q72"),
    "adl_eating_fim_current_val": ("様式23_1", "S35"),
    "adl_grooming_fim_current_val": ("様式23_1", "S36"),
    "adl_bathing_fim_current_val": ("様式23_1", "S37"),
    "adl_dressing_upper_fim_current_val": ("様式23_1", "S38"),
    "adl_dressing_lower_fim_current_val": ("様式23_1", "S39"),
    "adl_toileting_fim_current_val": ("様式23_1", "S40"),
    "adl_bladder_management_fim_current_val": ("様式23_1", "S41"),
    "adl_bowel_management_fim_current_val": ("様式23_1", "S42"),
    "adl_transfer_bed_chair_wc_fim_current_val": ("様式23_1", "S43"),
    "adl_transfer_toilet_fim_current_val": ("様式23_1", "S44"),
    "adl_transfer_tub_shower_fim_current_val": ("様式23_1", "S45"),
    "adl_locomotion_walk_walkingAids_wc_fim_current_val": ("様式23_1", "S46"),
    "adl_locomotion_stairs_fim_current_val": ("様式23_1", "S49"),
    "adl_comprehension_fim_current_val": ("様式23_1", "S51"),
    "adl_expression_fim_current_val": ("様式23_1", "S52"),
    "adl_social_interaction_fim_current_val": ("様式23_1", "S53"),
    "adl_problem_solving_fim_current_val": ("様式23_1", "S54"),
    "adl_memory_fim_current_val": ("様式23_1", "S55"),
    "nutrition_method_oral_meal_chk": ("様式23_1", "S61"),
    "func_respiratory_ventilator_chk": ("様式23_1", "T14"),
    "func_risk_smoking_chk": ("様式23_1", "T18"),
    "func_risk_family_history_chk": ("様式23_1", "T19"),
    "social_disability_certificate_physical_chk": ("様式23_1", "T70"),
    "social_disability_certificate_physical_txt": ("様式23_1", "T71"),
    "social_disability_certificate_physical_type_txt": ("様式23_1", "T72"),
    "signature_primary_doctor_txt": ("様式23_1", "T86"),
    "signature_ot_txt": ("様式23_1", "T87"),
    "signature_nurse_txt": ("様式23_1", "T88"),
    "signature_social_worker_txt": ("様式23_1", "T89"),
    "adl_eating_bi_start_val": ("様式23_1", "V35"),
    "adl_grooming_bi_start_val": ("様式23_1", "V36"),
    "adl_bathing_bi_start_val": ("様式23_1", "V37"),
    "adl_dressing_bi_start_val": ("様式23_1", "V38"),
    "adl_toileting_bi_start_val": ("様式23_1", "V40"),
    "adl_bladder_management_bi_start_val": ("様式23_1", "V41"),
    "adl_bowel_management_bi_start_val": ("様式23_1", "V42"),
    "adl_transfer_bi_start_val": ("様式23_1", "V43"),
    "adl_locomotion_walk_walkingAids_wc_bi_start_val": ("様式23_1", "V46"),
    "adl_locomotion_stairs_bi_start_val": ("様式23_1", "V49"),
    "nutrition_method_oral_supplement_chk": ("様式23_1", "V61"),
    "nutrition_status_assessment_malnutrition_risk_chk": ("様式23_1", "V63"),
    "nutrition_weight_chk": ("様式23_1", "W60"),
    "nutrition_required_energy_val": ("様式23_1", "W65"),
    "nutrition_total_intake_energy_val": ("様式23_1", "W66"),
    "social_disability_certificate_physical_rank_val": ("様式23_1", "Y72"),
    "func_contracture_deformity_chk": ("様式23_1", "Z13"),
    "func_muscle_weakness_chk": ("様式23_1", "Z14"),
    "func_motor_dysfunction_chk": ("様式23_1", "Z15"),
    "func_motor_muscle_tone_abnormality_chk": ("様式23_1", "Z17"),
    "func_sensory_dysfunction_chk": ("様式23_1", "Z18"),
    "func_speech_disorder_chk": ("様式23_1", "Z19"),
    "func_higher_brain_dysfunction_chk": ("様式23_1", "Z21"),
    "func_behavioral_psychiatric_disorder_chk": ("様式23_1", "Z22"),
    "func_disorientation_chk": ("様式23_1", "Z23"),
    "func_memory_disorder_chk": ("様式23_1", "Z24"),
    "func_developmental_disorder_chk": ("様式23_1", "Z25"),
    "func_basic_sitting_balance_chk": ("様式23_1", "Z29"),
    "func_basic_standing_balance_chk": ("様式23_1", "Z30"),
    "func_basic_other_chk": ("様式23_1", "Z31"),
    "goal_p_household_role_chk": ("様式23_2", "D10"),
    "goal_p_social_activity_chk": ("様式23_2", "D11"),
    "goal_p_hobby_chk": ("様式23_2", "D12"),
    "goal_a_bed_mobility_chk": ("様式23_2", "D13"),
    "goal_a_indoor_mobility_chk": ("様式23_2", "D16"),
    "goal_a_outdoor_mobility_chk": ("様式23_2", "D19"),
    "goal_p_residence_chk": ("様式23_2", "D2"),
    "goal_a_driving_chk": ("様式23_2", "D22"),
    "goal_a_public_transport_chk": ("様式23_2", "D25"),
    "goal_a_toileting_chk": ("様式23_2", "D28"),
    "goal_a_eating_chk": ("様式23_2", "D31"),
    "goal_a_grooming_chk": ("様式23_2", "D35"),
    "goal_a_dressing_chk": ("様式23_2", "D36"),
    "goal_a_bathing_chk": ("様式23_2", "D37"),
    "goal_p_return_to_work_chk": ("様式23_2", "D4"),
    "goal_a_housework_meal_chk": ("様式23_2", "D40"),
    "goal_a_writing_chk": ("様式23_2", "D42"),
    "goal_a_ict_chk": ("様式23_2", "D44"),
    "goal_a_communication_chk": ("様式23_2", "D46"),
    "goal_s_psychological_support_chk": ("様式23_2", "D52"),
    "goal_s_disability_acceptance_chk": ("様式23_2", "D53"),
    "goal_s_psychological_other_chk": ("様式23_2", "D54"),
    "goal_s_env_home_modification_chk": ("様式23_2", "D55"),
    "goal_s_env_assistive_device_chk": ("様式23_2", "D56"),
    "goal_s_env_social_security_chk": ("様式23_2", "D57"),
    "goal_s_env_care_insurance_chk": ("様式23_2", "D59"),
    "goal_s_env_disability_welfare_chk": ("様式23_2", "D62"),
    "goal_s_env_other_chk": ("様式23_2", "D64"),
    "goal_s_3rd_party_main_caregiver_chk": ("様式23_2", "D65"),
    "goal_s_3rd_party_family_structure_change_chk": ("様式23_2", "D66"),
    "goal_s_3rd_party_household_role_change_chk": ("様式23_2", "D67"),
    "goal_s_3rd_party_family_activity_change_chk": ("様式23_2", "D68"),
    "goal_p_schooling_chk": ("様式23_2", "D7"),
    "goal_a_bed_mobility_independent_chk": ("様式23_2", "E14"),
    "goal_a_bed_mobility_equipment_chk": ("様式23_2", "E15"),
    "goal_a_indoor_mobility_independent_chk": ("様式23_2", "E17"),
    "goal_a_indoor_mobility_equipment_chk": ("様式23_2", "E18"),
    "goal_a_outdoor_mobility_independent_chk": ("様式23_2", "E20"),
    "goal_a_outdoor_mobility_equipment_chk": ("様式23_2", "E21"),
    "goal_a_driving_independent_chk": ("様式23_2", "E23"),
    "goal_a_driving_modification_chk": ("様式23_2", "E24"),
    "goal_a_public_transport_independent_chk": ("様式23_2", "E26"),
    "goal_a_public_transport_type_chk": ("様式23_2", "E27"),
    "goal_a_toileting_independent_chk": ("様式23_2", "E29"),
    "goal_p_residence_home_type_slct": ("様式23_2", "E3"),
    "goal_a_toileting_type_chk": ("様式23_2", "E30"),
    "goal_a_eating_independent_chk": ("様式23_2", "E32"),
    "goal_a_eating_method_chopsticks_chk": ("様式23_2", "E33"),
    "goal_a_bathing_type_tub_chk": ("様式23_2", "E38"),
    "goal_a_bathing_assistance_body_washing_chk": ("様式23_2", "E39"),
    "goal_a_housework_meal_all_chk": ("様式23_2", "E41"),
    "goal_a_writing_independent_chk": ("様式23_2", "E43"),
    "goal_a_ict_independent_chk": ("様式23_2", "E45"),
    "goal_a_communication_independent_chk": ("様式23_2", "E47"),
    "goal_a_communication_device_chk": ("様式23_2", "E48"),
    "goal_p_return_to_work_status_current_job_chk": ("様式23_2", "E5"),
    "goal_s_env_social_security_physical_disability_cert_chk": ("様式23_2", "E58"),
    "goal_p_return_to_work_commute_change_chk": ("様式23_2", "E6"),
    "goal_s_env_care_insurance_outpatient_rehab_chk": ("様式23_2", "E60"),
    "goal_s_env_care_insurance_health_facility_chk": ("様式23_2", "E61"),
    "goal_s_env_disability_welfare_after_school_day_service_chk": ("様式23_2", "E63"),
    "goal_p_schooling_status_possible_chk": ("様式23_2", "E8"),
    "goal_p_schooling_destination_chk": ("様式23_2", "E9"),
    "goal_p_residence_home_type_detachedhouse_slct": ("様式23_2", "H3"),
    "goal_a_toileting_type_western_chk": ("様式23_2", "H30"),
    "goal_a_eating_method_fork_etc_chk": ("様式23_2", "H33"),
    "goal_a_grooming_independent_chk": ("様式23_2", "H35"),
    "goal_a_dressing_independent_chk": ("様式23_2", "H36"),
    "goal_a_bathing_independent_chk": ("様式23_2", "H37"),
    "goal_a_writing_independent_after_hand_change_chk": ("様式23_2", "H43"),
    "goal_p_schooling_status_needs_consideration_chk": ("様式23_2", "H8"),
    "goal_a_bed_mobility_assistance_chk": ("様式23_2", "I14"),
    "goal_a_indoor_mobility_assistance_chk": ("様式23_2", "I17"),
    "goal_a_outdoor_mobility_assistance_chk": ("様式23_2", "I20"),
    "goal_a_driving_assistance_chk": ("様式23_2", "I23"),
    "goal_a_driving_modification_txt": ("様式23_2", "I24"),
    "goal_a_public_transport_assistance_chk": ("様式23_2", "I26"),
    "goal_a_public_transport_type_txt": ("様式23_2", "I27"),
    "goal_a_toileting_assistance_chk": ("様式23_2", "I29"),
    "goal_a_eating_assistance_chk": ("様式23_2", "I32"),
    "goal_a_eating_diet_form_txt": ("様式23_2", "I34"),
    "goal_a_bathing_type_shower_chk": ("様式23_2", "I38"),
    "goal_a_bathing_assistance_transfer_chk": ("様式23_2", "I39"),
    "goal_a_housework_meal_not_performed_chk": ("様式23_2", "I41"),
    "goal_a_ict_assistance_chk": ("様式23_2", "I45"),
    "goal_a_communication_assistance_chk": ("様式23_2", "I47"),
    "goal_p_return_to_work_status_reassignment_chk": ("様式23_2", "I5"),
    "goal_s_psychological_support_txt": ("様式23_2", "I52"),
    "goal_s_disability_acceptance_txt": ("様式23_2", "I53"),
    "goal_s_psychological_other_txt": ("様式23_2", "I54"),
    "goal_s_env_social_security_disability_pension_chk": ("様式23_2", "I58"),
    "goal_s_env_care_insurance_home_rehab_chk": ("様式23_2", "I60"),
    "goal_s_env_care_insurance_nursing_home_chk": ("様式23_2", "I61"),
    "goal_s_env_disability_welfare_child_development_support_chk": ("様式23_2", "I63"),
    "goal_a_bed_mobility_environment_setup_chk": ("様式23_2", "J15"),
    "goal_s_env_home_modification_txt": ("様式23_2", "J55"),
    "goal_s_env_assistive_device_txt": ("様式23_2", "J56"),
    "goal_s_env_care_insurance_details_txt": ("様式23_2", "J59"),
    "goal_s_env_other_txt": ("様式23_2", "J64"),
    "goal_s_3rd_party_family_structure_change_txt": ("様式23_2", "J66"),
    "goal_p_schooling_destination_txt": ("様式23_2", "J9"),
    "goal_p_residence_home_type_apartment_slct": ("様式23_2", "K3"),
    "goal_a_toileting_type_japanese_chk": ("様式23_2", "K30"),
    "goal_s_3rd_party_main_caregiver_txt": ("様式23_2", "K65"),
    "goal_s_3rd_party_household_role_change_txt": ("様式23_2", "K67"),
    "goal_s_3rd_party_family_activity_change_txt": ("様式23_2", "K68"),
    "goal_a_indoor_mobility_not_performed_chk": ("様式23_2", "L17"),
    "goal_a_indoor_mobility_equipment_txt": ("様式23_2", "L18"),
    "goal_a_outdoor_mobility_not_performed_chk": ("様式23_2", "L20"),
    "goal_a_outdoor_mobility_equipment_txt": ("様式23_2", "L21"),
    "goal_a_driving_not_performed_chk": ("様式23_2", "L23"),
    "goal_a_public_transport_not_performed_chk": ("様式23_2", "L26"),
    "goal_a_toileting_assistance_clothing_chk": ("様式23_2", "L29"),
    "goal_a_eating_not_performed_chk": ("様式23_2", "L32"),
    "goal_a_eating_method_tube_feeding_chk": ("様式23_2", "L33"),
    "goal_a_grooming_assistance_chk": ("様式23_2", "L35"),
    "goal_a_dressing_assistance_chk": ("様式23_2", "L36"),
    "goal_a_bathing_assistance_chk": ("様式23_2", "L37"),
    "goal_a_communication_letter_board_chk": ("様式23_2", "L48"),
    "goal_a_bed_mobility_not_performed_chk": ("様式23_2", "M14"),
    "goal_a_housework_meal_partial_chk": ("様式23_2", "M41"),
    "goal_p_return_to_work_status_new_job_chk": ("様式23_2", "M5"),
    "goal_s_env_social_security_intractable_disease_cert_chk": ("様式23_2", "M58"),
    "goal_s_env_care_insurance_day_care_chk": ("様式23_2", "M60"),
    "goal_s_env_care_insurance_care_hospital_chk": ("様式23_2", "M61"),
    "goal_p_schooling_status_change_course_chk": ("様式23_2", "M8"),
    "goal_a_toileting_type_other_chk": ("様式23_2", "N30"),
    "goal_a_writing_other_chk": ("様式23_2", "N43"),
    "goal_p_residence_facility_chk": ("様式23_2", "O3"),
    "goal_a_toileting_assistance_wiping_chk": ("様式23_2", "P29"),
    "goal_a_communication_cooperation_chk": ("様式23_2", "P48"),
    "goal_p_return_to_work_status_not_possible_chk": ("様式23_2", "P5"),
    "goal_p_schooling_status_not_possible_chk": ("様式23_2", "P8"),
    "goal_p_schooling_commute_change_chk": ("様式23_2", "P9"),
    "goal_a_toileting_type_other_txt": ("様式23_2", "Q30"),
    "goal_s_env_care_insurance_home_nursing_chk": ("様式23_2", "Q60"),
    "goal_s_env_disability_welfare_life_care_chk": ("様式23_2", "Q63"),
    "goal_p_residence_other_chk": ("様式23_2", "R3"),
    "goal_a_housework_meal_partial_txt": ("様式23_2", "R41"),
    "goal_a_writing_other_txt": ("様式23_2", "R43"),
    "goal_s_env_care_insurance_other_chk": ("様式23_2", "R61"),
    "goal_p_return_to_work_status_other_chk": ("様式23_2", "S5"),
    "goal_s_env_social_security_other_chk": ("様式23_2", "S58"),
    "goal_p_schooling_status_other_chk": ("様式23_2", "S8"),
    "goal_a_toileting_assistance_catheter_chk": ("様式23_2", "T29"),
    "goal_p_residence_other_txt": ("様式23_2", "U3"),
    "goal_s_env_care_insurance_home_care_chk": ("様式23_2", "U60"),
    "goal_p_schooling_commute_change_txt": ("様式23_2", "U9"),
    "goal_p_return_to_work_status_other_txt": ("様式23_2", "V5"),
    "goal_s_env_social_security_other_txt": ("様式23_2", "V58"),
    "goal_s_env_care_insurance_other_txt": ("様式23_2", "V61"),
    "goal_s_env_disability_welfare_other_chk": ("様式23_2", "V63"),
    "goal_p_schooling_status_other_txt": ("様式23_2", "V8"),
}

def _get_cell_by_address(wb, sheet_name, cell_address):
    """シート名とセル座標からセルオブジェクトを取得する（結合セル対応）"""
    try:
        ws = wb[sheet_name]
        cell = ws[cell_address]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cell_ranges:
                if cell.coordinate in merged_range:
                    return merged_range.min_cell
        return cell
    except Exception as e:
        print(f"   [エラー] シート '{sheet_name}' またはセル '{cell_address}' の取得に失敗: {e}")
        return None


def get_cell_by_name(wb, name):
    """名前付き範囲からセルオブジェクトを取得する"""
    try:
        defined_name = wb.defined_names[name]
        # destinations は (sheetname, address) のタプルのジェネレータ
        dests = list(defined_name.destinations)
        if dests:
            sheetname, address = dests[0]
            # アドレスから '$' を取り除く
            cleaned_address = address.replace('$', '')
            # 範囲指定の場合 (例: A1:A5)、左上のセルを返す
            if ':' in cleaned_address:
                cleaned_address = cleaned_address.split(':')[0]
            return wb[sheetname][cleaned_address]
        return None
    except KeyError:
        print(f"   [警告] 名前付き範囲 '{name}' がExcelファイル内に見つかりません。")
        return None
    except Exception as e:
        print(f"   [エラー] 名前付き範囲 '{name}' の取得中にエラー: {e}")
        return None


def write_date_to_sheet(wb, date_value, base_key):
    """日付を年・月・日のセルに分割して書き込む"""
    try:
        sheet_name, year_coord = COLUMN_TO_CELL_COORDINATE_MAP[f"{base_key}_year_txt"]
        _, month_coord = COLUMN_TO_CELL_COORDINATE_MAP[f"{base_key}_month_txt"]
        _, day_coord = COLUMN_TO_CELL_COORDINATE_MAP[f"{base_key}_day_txt"]

        _get_cell_by_address(wb, sheet_name, year_coord).value = date_value.year
        _get_cell_by_address(wb, sheet_name, month_coord).value = date_value.month
        _get_cell_by_address(wb, sheet_name, day_coord).value = date_value.day
        print(f"   [成功] 日付書き込み: {base_key} -> {date_value}")
    except KeyError:
        # 年月日セルが定義されていない場合、単一セルに書き込みを試みる
        try:
            sheet_name, cell_address = COLUMN_TO_CELL_COORDINATE_MAP[f"{base_key}_date"]
            _get_cell_by_address(wb, sheet_name, cell_address).value = date_value.strftime("%Y-%m-%d")
            print(f"   [成功] 日付書き込み (単一セル): {base_key} -> {date_value}")
        except KeyError:
            print(f"   [警告] スキップ: {base_key} に対応する日付セルが見つかりません。")
    except Exception as e:
        print(f"   [エラー] 日付 '{base_key}' の書き込み中にエラー: {e}")


def create_plan_sheet(plan_data, liked_items: dict = None):
    """【最終版・座標指定方式】Excelに計画書を書き込む"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    liked_items = liked_items or {} # liked_itemsがNoneの場合に空の辞書をセット
    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        raise

    # 1. メインのデータ書き込み処理
    for db_col_name, (sheet_name, cell_address) in COLUMN_TO_CELL_COORDINATE_MAP.items():
        if "date" in db_col_name or "gender" in db_col_name:
            continue
        
        # 治療方針のキーは後で個別に処理するのでスキップ
        if db_col_name in ["header_therapy_pt_chk", "header_therapy_ot_chk", "header_therapy_st_chk"]:
            continue

        value = plan_data.get(db_col_name)
        if value is None or value == "":
            continue

        target_cell = _get_cell_by_address(wb, sheet_name, cell_address)

        if not target_cell:
            print(f"   [警告] スキップ: セル '{sheet_name}!{cell_address}' が見つかりません。")
            continue

        try:
            print(f"   [成功] 書き込み中: '{sheet_name}!{cell_address}' に値を設定します。")

            # このループでは、単純なブール値を持つチェックボックスと、その他のテキスト/数値のみを処理します。
            # 住宅関連のキーはここでは処理されません。
            if isinstance(value, bool) or value in (1, 0):
                target_cell.value = "☑" if value else "☐"
            elif 'goal_p_residence_home_type' not in db_col_name: # 住宅関連の特殊キーをここで除外
                # --- ▼▼▼ いいね情報を追記する処理を追加 ▼▼▼ ---
                liked_models = liked_items.get(db_col_name, []) # liked_itemsからリストを取得
                marks = ""
                if 'general' in liked_models:
                    marks += " [G]"
                if 'specialized' in liked_models:
                    marks += " [S]"
                if marks:
                    value = f"{value}{marks}"
                # --- ▲▲▲ いいね情報を追記する処理ここまで ▲▲▲ ---
                target_cell.value = value

        except Exception as e:
            print(f"   [エラー] '{sheet_name}!{cell_address}' の書き込み中にエラー: {e}")

    # 2. 特殊処理 (日付、性別、住宅種別)
    # 日付処理
    for key in ["header_evaluation_date", "header_onset_date", "header_rehab_start_date", "signature_explanation_date"]:
        date_value = plan_data.get(key)
        if isinstance(date_value, date):
            base_key = key.replace("_date", "")
            write_date_to_sheet(wb, date_value, base_key)

    # 性別処理
    try:
        gender = plan_data.get("gender")
        male_cell = _get_cell_by_address(wb, "様式23_1", "V3")
        female_cell = _get_cell_by_address(wb, "様式23_1", "X3")
        if male_cell and female_cell:
            font_selected = Font(size=13, bold=False)
            font_unselected = Font(size=11, bold=False)
            if gender == "男":
                male_cell.value, male_cell.font = "㊚", font_selected
                female_cell.value, female_cell.font = "女", font_unselected
                print("   [成功] 性別を '男' に設定しました。")
            elif gender == "女":
                male_cell.value, male_cell.font = "男", font_unselected
                female_cell.value, female_cell.font = "㊛", font_selected
                print("   [成功] 性別を '女' に設定しました。")
            else:
                male_cell.value, male_cell.font = "男", font_unselected
                female_cell.value, female_cell.font = "女", font_unselected
    except Exception as e:
        print(f"   [エラー] 性別の特殊処理中にエラー: {e}")

    # 治療方針のチェックボックスを名前付き範囲で処理
    try:
        pt_checked = plan_data.get("header_therapy_pt_chk", False)
        ot_checked = plan_data.get("header_therapy_ot_chk", False)
        st_checked = plan_data.get("header_therapy_st_chk", False)

        pt_cell = get_cell_by_name(wb, "header_therapy_pt_chk")
        ot_cell = get_cell_by_name(wb, "header_therapy_ot_chk")
        st_cell = get_cell_by_name(wb, "header_therapy_st_chk")

        if pt_cell:
            pt_cell.value = "☑" if pt_checked else "☐"
        if ot_cell:
            ot_cell.value = "☑" if ot_checked else "☐"
        if st_cell:
            st_cell.value = "☑" if st_checked else "☐"
        
        print("   [成功] 治療方針のチェックボックスを名前付き範囲で設定しました。")
    except Exception as e:
        print(f"   [エラー] 治療方針のチェックボックス処理中にエラー: {e}")


    # 住宅種別処理 (専用ブロック)
    try:
        residence_type = plan_data.get("goal_p_residence_slct")
        
        # マップから各チェックボックスのセル情報を取得
        home_key = "goal_p_residence_home_type_slct"
        detached_key = "goal_p_residence_home_type_detachedhouse_slct"
        apartment_key = "goal_p_residence_home_type_apartment_slct"

        # 自宅
        if home_key in COLUMN_TO_CELL_COORDINATE_MAP:
            sheet, cell = COLUMN_TO_CELL_COORDINATE_MAP[home_key]
            _get_cell_by_address(wb, sheet, cell).value = "☑" if residence_type in ["home_detached", "home_apartment"] else "☐"

        # 戸建
        if detached_key in COLUMN_TO_CELL_COORDINATE_MAP:
            sheet, cell = COLUMN_TO_CELL_COORDINATE_MAP[detached_key]
            _get_cell_by_address(wb, sheet, cell).value = "☑" if residence_type == "home_detached" else "☐"

        # マンション
        if apartment_key in COLUMN_TO_CELL_COORDINATE_MAP:
            sheet, cell = COLUMN_TO_CELL_COORDINATE_MAP[apartment_key]
            _get_cell_by_address(wb, sheet, cell).value = "☑" if residence_type == "home_apartment" else "☐"
        
        print(f"   [情報] 住宅種別 '{residence_type}' のチェックボックスを処理しました。")
    except Exception as e:
        print(f"   [エラー] 住宅種別の特殊処理中にエラー: {e}")

    # 3. ファイルの保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_patient_name = "".join(c for c in plan_data.get("name", "NoName") if c.isalnum())
    output_filename = f"RehabPlan_{safe_patient_name}_{timestamp}.xlsx"
    output_filepath = os.path.join(OUTPUT_DIR, output_filename)

    wb.save(output_filepath)
    print(f"\n計画書を {output_filepath} に保存しました。")

    return output_filepath
